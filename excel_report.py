#!/usr/bin/env python3
"""
excel_report.py — 월간 경영 보고서 엑셀 생성 모듈
기존 엑셀 양식과 동일한 포맷 (회색 헤더, 노란 입력셀, 파란 글씨, 수식 등)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === 원본 양식과 동일한 스타일 ===
F_TITLE     = Font(name='맑은 고딕', size=9, bold=True, color='000000')
F_SECTION   = Font(name='맑은 고딕', size=9, bold=True, color='000000')
F_NORMAL    = Font(name='맑은 고딕', size=9)
F_CATEGORY  = Font(name='맑은 고딕', size=8)
F_BLUE      = Font(name='맑은 고딕', size=9, color='0000FF')
F_GRAY      = Font(name='맑은 고딕', size=9, color='7F7F7F')
F_BOLD      = Font(name='맑은 고딕', size=9, bold=True)
F_GRAY_BOLD = Font(name='맑은 고딕', size=9, bold=True, color='7F7F7F')
F_RED_BOLD  = Font(name='맑은 고딕', size=9, bold=True, color='FF0000')
F_NOTE      = Font(name='맑은 고딕', size=7)
F_UNIT      = Font(name='맑은 고딕', size=9)

BG_SECTION  = PatternFill('solid', fgColor='D8D8D8')
BG_TBL_HEAD = PatternFill('solid', fgColor='D9D9D9')
BG_YELLOW   = PatternFill('solid', fgColor='FFFFCC')
BG_DATA     = PatternFill('solid', fgColor='F2F2F2')
BG_SUBTOTAL = PatternFill('solid', fgColor='D8D8D8')

AL_C  = Alignment(horizontal='center', vertical='center')
AL_R  = Alignment(horizontal='right', vertical='center')
AL_L  = Alignment(horizontal='left', vertical='center')
AL_LW = Alignment(horizontal='left', vertical='center', wrap_text=True)

_thin = Side(style='thin', color='C0C0C0')
BDR = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

NUM = '#,##0'
NUM_NEG = '#,##0;[Red]-#,##0;"-"'


def _row_cells(ws, r, col_range):
    """행의 지정 컬럼들에 테두리 적용"""
    for c in col_range:
        ws.cell(row=r, column=c).border = BDR


def generate_excel_report(data, output_path):
    """
    기존 엑셀 양식과 동일한 월간 경영 보고서 생성
    
    data 필수 키:
        year, month, company_name,
        revenue_detail, expense_detail,
        total_rev, total_opex, total_etc, total_invest, net_profit,
        key_points: [{'icon': str, 'text': str}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "월간보고서"
    
    year = data['year']
    month = data['month']
    company = data.get('company_name', '')
    
    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    prev_label = f"{prev_y}년 {prev_m}월 "
    
    total_rev = data.get('total_rev', 0)
    total_opex = data.get('total_opex', 0)
    total_etc = data.get('total_etc', 0)
    total_invest = data.get('total_invest', 0)
    
    # 컬럼 너비 (원본과 동일)
    for c, w in {1: 3, 2: 8, 3: 25, 4: 14, 5: 14, 6: 14, 7: 14, 8: 30}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    
    COLS = range(1, 9)  # A~H
    r = 2
    
    # =====================================================================
    # 제목 + 이슈
    # =====================================================================
    ws.cell(row=r, column=3, value=f"{year}년 {month}월 ").font = F_TITLE
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    ws.cell(row=r, column=3, value="■  주요이슈 및 특이사항").font = F_SECTION
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    points = data.get('key_points', [])
    for pt in points[:4]:
        text = pt.get('text', pt) if isinstance(pt, dict) else str(pt)
        icon = pt.get('icon', '•') if isinstance(pt, dict) else '•'
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.cell(row=r, column=3, value=f"  {icon} {text}").font = F_NORMAL
        r += 1
    r += 1
    
    # =====================================================================
    # 1. 사업자 총 매출
    # =====================================================================
    for c in range(3, 9):
        ws.cell(row=r, column=c).fill = BG_SECTION
    ws.cell(row=r, column=3, value="1. 사업자 총 매출 ").font = F_SECTION
    ws.cell(row=r, column=3).fill = BG_SECTION
    cell = ws.cell(row=r, column=8, value=total_rev)
    cell.font = F_BOLD; cell.number_format = NUM; cell.fill = BG_SECTION; cell.alignment = AL_R
    ws.row_dimensions[r].height = 22.5
    r += 2
    
    # =====================================================================
    # 2. 브랜드별 수익현황
    # =====================================================================
    for c in range(3, 9):
        ws.cell(row=r, column=c).fill = BG_SECTION
    ws.cell(row=r, column=3, value="2. 브랜드별 수익현황").font = F_SECTION
    ws.cell(row=r, column=3).fill = BG_SECTION
    ws.cell(row=r, column=8, value='(단위:원)').font = F_UNIT
    ws.cell(row=r, column=8).fill = BG_SECTION
    ws.cell(row=r, column=8).alignment = AL_R
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    # 합계 표시
    cell = ws.cell(row=r, column=8, value=total_rev)
    cell.font = F_BOLD; cell.number_format = NUM; cell.alignment = AL_R
    r += 1
    
    # 테이블 헤더
    th_data = [(2, '구분'), (4, f'{year}년 {month}월'), (5, company), (7, prev_label), (8, '비고')]
    for c in COLS:
        ws.cell(row=r, column=c).border = BDR
        ws.cell(row=r, column=c).alignment = AL_C
        if c in (5, 6):
            ws.cell(row=r, column=c).fill = BG_YELLOW
            ws.cell(row=r, column=c).font = F_BLUE
        else:
            ws.cell(row=r, column=c).fill = BG_TBL_HEAD
            ws.cell(row=r, column=c).font = F_NORMAL
    for c, v in th_data:
        ws.cell(row=r, column=c, value=v)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    # 매출 데이터
    rev_detail = data.get('revenue_detail', {})
    rev_start = r
    
    for sub, amt in rev_detail.items():
        _row_cells(ws, r, COLS)
        ws.cell(row=r, column=2).fill = BG_YELLOW
        ws.cell(row=r, column=3, value=sub).font = F_NORMAL
        ws.cell(row=r, column=3).fill = BG_YELLOW
        ws.cell(row=r, column=3).border = BDR
        
        cell_d = ws.cell(row=r, column=4, value=int(amt))
        cell_d.font = F_NORMAL; cell_d.number_format = NUM; cell_d.alignment = AL_R; cell_d.fill = BG_YELLOW
        
        cell_e = ws.cell(row=r, column=5, value=int(amt))
        cell_e.font = F_BLUE; cell_e.fill = BG_YELLOW; cell_e.number_format = NUM; cell_e.alignment = AL_R
        
        ws.cell(row=r, column=6).fill = BG_YELLOW
        ws.cell(row=r, column=7).font = F_GRAY
        ws.cell(row=r, column=8).font = F_NOTE
        r += 1
    
    # 매출 합계 행
    _row_cells(ws, r, COLS)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value='합계').font = F_BOLD
    ws.cell(row=r, column=2).alignment = AL_C
    
    ws.cell(row=r, column=4, value=f'=SUM(D{rev_start}:D{r-1})').font = F_BOLD
    ws.cell(row=r, column=4).number_format = NUM; ws.cell(row=r, column=4).alignment = AL_R
    
    ws.cell(row=r, column=5, value=f'=SUM(E{rev_start}:E{r-1})').font = F_BLUE
    ws.cell(row=r, column=5).fill = BG_YELLOW; ws.cell(row=r, column=5).number_format = NUM
    ws.cell(row=r, column=5).alignment = AL_R
    ws.cell(row=r, column=6).fill = BG_YELLOW
    
    rev_total_row = r
    r += 2
    
    # =====================================================================
    # 3. 판관비 지출현황
    # =====================================================================
    for c in range(3, 9):
        ws.cell(row=r, column=c).fill = BG_SECTION
    ws.cell(row=r, column=3, value="3. 판관비 지출현황 ").font = F_SECTION
    ws.cell(row=r, column=3).fill = BG_SECTION
    ws.cell(row=r, column=8, value='(단위:원)').font = F_UNIT
    ws.cell(row=r, column=8).fill = BG_SECTION; ws.cell(row=r, column=8).alignment = AL_R
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    cell = ws.cell(row=r, column=8, value=total_opex)
    cell.font = F_BOLD; cell.number_format = NUM; cell.alignment = AL_R
    r += 1
    
    # 헤더
    th_data2 = [(3, '구분'), (4, f'{year}년 {month}월'), (5, company), (7, prev_label), (8, '비고')]
    for c in COLS:
        ws.cell(row=r, column=c).border = BDR; ws.cell(row=r, column=c).alignment = AL_C
        if c in (5, 6):
            ws.cell(row=r, column=c).fill = BG_YELLOW; ws.cell(row=r, column=c).font = F_BLUE
        else:
            ws.cell(row=r, column=c).fill = BG_TBL_HEAD; ws.cell(row=r, column=c).font = F_NORMAL
    for c, v in th_data2:
        ws.cell(row=r, column=c, value=v)
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    # 판관비 데이터
    exp_detail = data.get('expense_detail', {})
    exp_start = r
    
    for sub, amt in exp_detail.items():
        _row_cells(ws, r, COLS)
        ws.cell(row=r, column=2).fill = BG_DATA
        
        ws.cell(row=r, column=3, value=f' {sub} ').font = F_NORMAL
        ws.cell(row=r, column=3).fill = BG_DATA
        
        cell_d = ws.cell(row=r, column=4, value=int(amt))
        cell_d.font = F_NORMAL; cell_d.number_format = NUM; cell_d.alignment = AL_R; cell_d.fill = BG_DATA
        
        cell_e = ws.cell(row=r, column=5, value=int(amt))
        cell_e.font = F_BLUE; cell_e.fill = BG_YELLOW; cell_e.number_format = NUM; cell_e.alignment = AL_R
        
        ws.cell(row=r, column=6).fill = BG_YELLOW
        ws.cell(row=r, column=7).font = F_GRAY; ws.cell(row=r, column=7).fill = BG_DATA
        ws.cell(row=r, column=8).font = F_NOTE; ws.cell(row=r, column=8).fill = BG_DATA
        r += 1
    
    # 판관비 합계
    _row_cells(ws, r, COLS)
    ws.cell(row=r, column=3, value='총합계').font = F_BOLD
    ws.cell(row=r, column=4, value=f'=SUM(D{exp_start}:D{r-1})').font = F_BOLD
    ws.cell(row=r, column=4).number_format = NUM; ws.cell(row=r, column=4).alignment = AL_R
    ws.cell(row=r, column=5, value=f'=SUM(E{exp_start}:E{r-1})').font = F_BLUE
    ws.cell(row=r, column=5).fill = BG_YELLOW; ws.cell(row=r, column=5).number_format = NUM
    ws.cell(row=r, column=5).alignment = AL_R
    ws.cell(row=r, column=6).fill = BG_YELLOW
    
    exp_total_row = r
    r += 1
    
    # 매출 - 판관비
    cell = ws.cell(row=r, column=8, value=f'=D{rev_total_row}-D{exp_total_row}')
    cell.font = F_RED_BOLD; cell.number_format = NUM_NEG; cell.alignment = AL_R
    r += 1
    
    # =====================================================================
    # 4. 기타비용 지출현황
    # =====================================================================
    etc_detail = data.get('etc_detail', {})
    invest_detail = data.get('invest_detail', {})
    all_etc = {**etc_detail, **invest_detail}
    etc_total_val = total_etc + total_invest
    
    for c in range(3, 9):
        ws.cell(row=r, column=c).fill = BG_SECTION
    ws.cell(row=r, column=3, value="4. 기타비용 지출현황 ").font = F_SECTION
    ws.cell(row=r, column=3).fill = BG_SECTION
    ws.cell(row=r, column=8, value='(단위:원)').font = F_UNIT
    ws.cell(row=r, column=8).fill = BG_SECTION; ws.cell(row=r, column=8).alignment = AL_R
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    cell = ws.cell(row=r, column=8, value=etc_total_val)
    cell.font = F_BOLD; cell.number_format = NUM; cell.alignment = AL_R
    r += 1
    
    # 헤더
    for c in COLS:
        ws.cell(row=r, column=c).border = BDR; ws.cell(row=r, column=c).alignment = AL_C
        if c in (5, 6):
            ws.cell(row=r, column=c).fill = BG_YELLOW; ws.cell(row=r, column=c).font = F_BLUE
        else:
            ws.cell(row=r, column=c).fill = BG_TBL_HEAD; ws.cell(row=r, column=c).font = F_NORMAL
    for c, v in th_data2:
        ws.cell(row=r, column=c, value=v)
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    etc_start = r
    
    if all_etc:
        for sub, amt in all_etc.items():
            _row_cells(ws, r, COLS)
            ws.cell(row=r, column=3, value=f' {sub}').font = F_NORMAL; ws.cell(row=r, column=3).border = BDR
            cell_d = ws.cell(row=r, column=4, value=int(amt))
            cell_d.font = F_NORMAL; cell_d.number_format = NUM; cell_d.alignment = AL_R
            cell_e = ws.cell(row=r, column=5, value=int(amt))
            cell_e.font = F_BLUE; cell_e.fill = BG_YELLOW; cell_e.number_format = NUM; cell_e.alignment = AL_R
            ws.cell(row=r, column=6).fill = BG_YELLOW
            ws.cell(row=r, column=7).font = F_GRAY
            ws.cell(row=r, column=8).font = F_NOTE
            ws.row_dimensions[r].height = 22.5
            r += 1
    else:
        _row_cells(ws, r, COLS)
        ws.cell(row=r, column=3, value=' (해당 없음)').font = F_NORMAL
        ws.cell(row=r, column=4, value=0).number_format = NUM
        ws.cell(row=r, column=5).fill = BG_YELLOW
        ws.cell(row=r, column=6).fill = BG_YELLOW
        r += 1
    
    # 기타 합계
    _row_cells(ws, r, COLS)
    ws.cell(row=r, column=3, value='총합계').font = F_BOLD
    ws.cell(row=r, column=4, value=f'=SUM(D{etc_start}:D{r-1})').font = F_BOLD
    ws.cell(row=r, column=4).number_format = NUM; ws.cell(row=r, column=4).alignment = AL_R
    ws.cell(row=r, column=5).fill = BG_YELLOW
    ws.cell(row=r, column=6).fill = BG_YELLOW
    etc_total_row = r
    r += 2
    
    # =====================================================================
    # 5. 최종 결산
    # =====================================================================
    for c in range(3, 9):
        ws.cell(row=r, column=c).fill = BG_SECTION
    ws.cell(row=r, column=3, value="5. 최종 결산 ").font = F_SECTION
    ws.cell(row=r, column=3).fill = BG_SECTION
    ws.cell(row=r, column=8, value='(단위:원)').font = F_UNIT
    ws.cell(row=r, column=8).fill = BG_SECTION; ws.cell(row=r, column=8).alignment = AL_R
    ws.row_dimensions[r].height = 22.5
    r += 1
    
    # 최종 행
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    ws.cell(row=r, column=3, value="수익 - 판관비 - 기타지출 ").font = F_BOLD
    ws.cell(row=r, column=3).fill = BG_YELLOW
    
    result_formula = f'=D{rev_total_row}-D{exp_total_row}-D{etc_total_row}'
    result_cell = ws.cell(row=r, column=8, value=result_formula)
    result_cell.font = F_RED_BOLD
    result_cell.fill = BG_YELLOW
    result_cell.number_format = NUM_NEG
    result_cell.alignment = AL_R
    
    for c in range(3, 9):
        ws.cell(row=r, column=c).border = BDR
    
    # 인쇄 설정
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    
    wb.save(output_path)
    return output_path
